import asyncio
import csv
import io
import time as time_module
import uuid
from collections.abc import AsyncIterator, Callable, Sized
from datetime import UTC, datetime
from typing import cast

import numpy as np
import torch
from matplotlib.figure import Figure
from openpyxl import Workbook
from sqlalchemy.exc import SQLAlchemyError
from torch import nn
from torch.utils.data import DataLoader, Dataset

from app.models.lstm import LstmTrainingMetrics, LstmTrainRequest
from app.models.tables.databaseTables import LstmResult
from app.repositories import lstmRepo
from app.utils.plot import get_chinese_font

ProgressCallback = Callable[[dict[str, object]], None]

_LSTM_PARAMETER_LABELS = {
    "n_samples": "样本数量(n_samples)",
    "time_step": "时间步长(time_step)",
    "trend_slope": "趋势斜率(trend_slope)",
    "primary_amplitude": "主周期振幅(primary_amplitude)",
    "primary_frequency": "主周期频率(primary_frequency)",
    "secondary_amplitude": "次周期振幅(secondary_amplitude)",
    "secondary_frequency": "次周期频率(secondary_frequency)",
    "noise_std": "噪声标准差(noise_std)",
    "seed": "随机种子(seed)",
    "forecast_horizon": "预测步数(forecast_horizon)",
    "sequence_length": "输入窗口长度(sequence_length)",
    "hidden_size": "隐藏层维度(hidden_size)",
    "num_layers": "LSTM层数(num_layers)",
    "dropout": "丢弃率(dropout)",
    "epochs": "训练轮数(epochs)",
    "batch_size": "批大小(batch_size)",
    "learning_rate": "学习率(learning_rate)",
    "train_ratio": "训练集比例(train_ratio)",
    "device": "计算设备(device)",
}


class TimeSeriesDataset(Dataset):
    def __init__(self, features, data, sequence_length, horizon):
        self.features = features
        self.data = data
        self.sequence_length = sequence_length
        self.horizon = horizon

    def __len__(self):
        return len(self.data) - self.sequence_length - self.horizon + 1

    def __getitem__(self, index):
        target_start = index + self.sequence_length
        target_end = target_start + self.horizon
        history = self.features[index:target_start]
        future_time_features = self.features[target_start:target_end, 1:]
        changes = self.data[target_start:target_end] - self.data[target_start - 1]
        return (
            torch.from_numpy(history),
            torch.from_numpy(future_time_features),
            torch.from_numpy(changes.astype(np.float32)),
        )


class LstmPredictor(nn.Module):
    def __init__(
        self,
        input_size: int,
        hidden_size: int,
        num_layers: int,
        dropout: float,
        future_feature_size: int,
    ):
        super().__init__()
        self.lstm = nn.LSTM(
            input_size=input_size,
            hidden_size=hidden_size,
            num_layers=num_layers,
            batch_first=True,
            dropout=dropout if num_layers > 1 else 0,
        )
        self.head = nn.Sequential(
            nn.Linear(hidden_size + future_feature_size, hidden_size),
            nn.ReLU(),
            nn.Linear(hidden_size, 1),
        )

    def forward(self, history, future_time_features):
        output, _ = self.lstm(history)
        context = output[:, -1, :]
        horizon = future_time_features.size(1)
        repeated_context = context.unsqueeze(1).expand(-1, horizon, -1)
        decoder_input = torch.cat((repeated_context, future_time_features), dim=-1)
        return self.head(decoder_input).squeeze(-1)


_training_semaphore = asyncio.Semaphore(1)


def _build_time_features(time_values: np.ndarray, time_scale: float) -> np.ndarray:
    return np.column_stack(
        (
            time_values / time_scale,
            np.sin(time_values),
            np.cos(time_values),
        )
    ).astype(np.float32)


def _build_features(
    normalized_data: np.ndarray,
    time_values: np.ndarray,
    time_scale: float,
) -> np.ndarray:
    return np.column_stack(
        (normalized_data, _build_time_features(time_values, time_scale))
    ).astype(np.float32)


def _series_values(
    time_values: np.ndarray,
    request: LstmTrainRequest,
    noise: np.ndarray,
) -> np.ndarray:
    params = request.synthesis
    trend = params.trend_slope * time_values
    primary = params.primary_amplitude * np.sin(params.primary_frequency * time_values)
    secondary = params.secondary_amplitude * np.sin(
        params.secondary_frequency * time_values
    )
    return trend + primary + secondary + noise


def _resolve_device(requested_device: str) -> torch.device:
    if requested_device == "cuda":
        if not torch.cuda.is_available():
            raise ValueError("当前服务环境不支持 CUDA")
        return torch.device("cuda")
    if requested_device == "auto" and torch.cuda.is_available():
        return torch.device("cuda")
    return torch.device("cpu")


def _train_model(
    model: LstmPredictor,
    train_loader: DataLoader,
    validation_loader: DataLoader,
    request: LstmTrainRequest,
    device: torch.device,
    progress_callback: ProgressCallback | None = None,
) -> tuple[list[float], list[float]]:
    criterion = nn.MSELoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=request.training.learning_rate)

    train_losses: list[float] = []
    validation_losses: list[float] = []
    for epoch in range(1, request.training.epochs + 1):
        model.train()
        train_total = 0.0
        for history, future_time, target in train_loader:
            history = history.to(device)
            future_time = future_time.to(device)
            target = target.to(device)
            optimizer.zero_grad()
            prediction = model(history, future_time)
            loss = criterion(prediction, target)
            loss.backward()
            optimizer.step()
            train_total += loss.item() * history.size(0)
        train_sample_count = len(cast(Sized, train_loader.dataset))
        train_losses.append(train_total / train_sample_count)

        model.eval()
        validation_total = 0.0
        with torch.no_grad():
            for history, future_time, target in validation_loader:
                history = history.to(device)
                future_time = future_time.to(device)
                target = target.to(device)
                prediction = model(history, future_time)
                validation_total += criterion(prediction, target).item() * history.size(
                    0
                )
        validation_sample_count = len(cast(Sized, validation_loader.dataset))
        validation_losses.append(validation_total / validation_sample_count)
        if progress_callback is not None:
            progress_callback(
                {
                    "type": "epoch",
                    "epoch": epoch,
                    "total_epochs": request.training.epochs,
                    "train_loss": train_losses[-1],
                    "validation_loss": validation_losses[-1],
                }
            )

    return train_losses, validation_losses


@torch.no_grad()
def _predict_future(
    model: LstmPredictor,
    features: np.ndarray,
    normalized_data: np.ndarray,
    future_time: np.ndarray,
    time_scale: float,
    sequence_length: int,
    device: torch.device,
) -> np.ndarray:
    model.eval()
    history = torch.from_numpy(features[-sequence_length:]).unsqueeze(0).to(device)
    future_features = (
        torch.from_numpy(_build_time_features(future_time, time_scale))
        .unsqueeze(0)
        .to(device)
    )
    changes = model(history, future_features).squeeze(0).cpu().numpy()
    return normalized_data[-1] + changes


def _render_image(
    observed_time: np.ndarray,
    observed: np.ndarray,
    future_time: np.ndarray,
    forecast: np.ndarray,
    expected: np.ndarray,
    train_losses: list[float],
    validation_losses: list[float],
) -> bytes:
    chinese_font = get_chinese_font()
    figure = Figure(figsize=(14, 8), constrained_layout=True, facecolor="#f4f6f8")
    loss_axis, forecast_axis = figure.subplots(2, 1)
    figure.suptitle(
        "LSTM 时间序列预测结果",
        fontproperties=chinese_font,
        fontsize=18,
        fontweight="bold",
    )

    epochs = np.arange(1, len(train_losses) + 1)
    loss_axis.plot(
        epochs,
        train_losses,
        label="训练损失",
        color="#2563eb",
        linewidth=2,
        marker="o",
        markersize=4,
    )
    loss_axis.plot(
        epochs,
        validation_losses,
        label="验证损失",
        color="#f97316",
        linewidth=2,
        marker="o",
        markersize=4,
    )
    loss_axis.set_title("训练损失变化", fontproperties=chinese_font, fontsize=13)
    loss_axis.set_xlabel("训练轮次", fontproperties=chinese_font)
    loss_axis.set_ylabel("均方误差损失", fontproperties=chinese_font)
    loss_axis.grid(alpha=0.2, linestyle="--")
    loss_axis.legend(prop=chinese_font, frameon=False)

    forecast_axis.plot(
        observed_time,
        observed,
        label="历史观测值",
        color="#475569",
        alpha=0.8,
    )
    forecast_axis.plot(
        future_time,
        expected,
        label="未来真实值",
        color="#16a34a",
        linestyle="--",
        linewidth=2,
    )
    forecast_axis.plot(
        future_time,
        forecast,
        label="模型预测值",
        color="#dc2626",
        linewidth=2.2,
    )
    forecast_axis.axvline(
        observed_time[-1],
        color="#64748b",
        linestyle=":",
        label="预测起点",
    )
    forecast_axis.set_title("多步预测结果", fontproperties=chinese_font, fontsize=13)
    forecast_axis.set_xlabel("时间", fontproperties=chinese_font)
    forecast_axis.set_ylabel("数值", fontproperties=chinese_font)
    forecast_axis.grid(alpha=0.2, linestyle="--")
    forecast_axis.legend(prop=chinese_font, frameon=False, ncol=2)

    for axis in (loss_axis, forecast_axis):
        axis.set_facecolor("white")
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)

    output = io.BytesIO()
    figure.savefig(output, format="png", dpi=150)
    return output.getvalue()


def _build_csv(
    observed_time: np.ndarray,
    observed: np.ndarray,
    future_time: np.ndarray,
    forecast: np.ndarray,
    expected: np.ndarray,
) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(("步骤(step)", "时间(time)", "序列(series)", "数值(value)"))
    for step, (time_value, value) in enumerate(
        zip(observed_time, observed, strict=True)
    ):
        writer.writerow((step, float(time_value), "观测值(observed)", float(value)))
    offset = len(observed)
    for step, (time_value, predicted, actual) in enumerate(
        zip(future_time, forecast, expected, strict=True), start=offset
    ):
        writer.writerow((step, float(time_value), "预测值(forecast)", float(predicted)))
        writer.writerow((step, float(time_value), "真实值(expected)", float(actual)))
    return output.getvalue().encode("utf-8-sig")


def _build_excel(
    observed_time: np.ndarray,
    observed: np.ndarray,
    future_time: np.ndarray,
    forecast: np.ndarray,
    expected: np.ndarray,
    train_losses: list[float],
    validation_losses: list[float],
    dataset_params: dict,
    model_params: dict,
    training_params: dict,
) -> bytes:
    workbook = Workbook()
    series_sheet = workbook.active
    if series_sheet is None:
        series_sheet = workbook.create_sheet()
    series_sheet.title = "数据序列(series)"
    series_sheet.append(("步骤(step)", "时间(time)", "序列(series)", "数值(value)"))
    for step, (time_value, value) in enumerate(
        zip(observed_time, observed, strict=True)
    ):
        series_sheet.append((step, float(time_value), "观测值(observed)", float(value)))
    offset = len(observed)
    for step, (time_value, predicted, actual) in enumerate(
        zip(future_time, forecast, expected, strict=True), start=offset
    ):
        series_sheet.append(
            (step, float(time_value), "预测值(forecast)", float(predicted))
        )
        series_sheet.append(
            (step, float(time_value), "真实值(expected)", float(actual))
        )

    loss_sheet = workbook.create_sheet("损失(losses)")
    loss_sheet.append(
        ("轮次(epoch)", "训练损失(train_loss)", "验证损失(validation_loss)")
    )
    for epoch, (train_loss, validation_loss) in enumerate(
        zip(train_losses, validation_losses, strict=True), start=1
    ):
        loss_sheet.append((epoch, train_loss, validation_loss))

    params_sheet = workbook.create_sheet("参数(parameters)")
    params_sheet.append(("分组(group)", "参数名(name)", "参数值(value)"))
    request_data = {
        "数据集(dataset)": dataset_params,
        "模型(model)": model_params,
        "训练(training)": training_params,
    }
    for group, values in request_data.items():
        for name, value in values.items():
            params_sheet.append((group, _LSTM_PARAMETER_LABELS.get(name, name), value))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def get_result(
    result_id: uuid.UUID,
    user_id: uuid.UUID,
) -> LstmResult | None:
    return lstmRepo.select_result_by_id_and_user_id(result_id, user_id)


def render_result_image(result: LstmResult) -> bytes:
    return _render_image(
        np.asarray(result.observed_time, dtype=float),
        np.asarray(result.observed_data, dtype=float),
        np.asarray(result.future_time, dtype=float),
        np.asarray(result.forecast_data, dtype=float),
        np.asarray(result.expected_data, dtype=float),
        result.train_losses,
        result.validation_losses,
    )


def export_result_csv(result: LstmResult) -> bytes:
    return _build_csv(
        np.asarray(result.observed_time, dtype=float),
        np.asarray(result.observed_data, dtype=float),
        np.asarray(result.future_time, dtype=float),
        np.asarray(result.forecast_data, dtype=float),
        np.asarray(result.expected_data, dtype=float),
    )


def export_result_excel(result: LstmResult) -> bytes:
    return _build_excel(
        np.asarray(result.observed_time, dtype=float),
        np.asarray(result.observed_data, dtype=float),
        np.asarray(result.future_time, dtype=float),
        np.asarray(result.forecast_data, dtype=float),
        np.asarray(result.expected_data, dtype=float),
        result.train_losses,
        result.validation_losses,
        result.dataset_params,
        result.model_params,
        result.training_params,
    )


def _train_and_store_sync(
    request: LstmTrainRequest,
    user_id: uuid.UUID,
    progress_callback: ProgressCallback | None = None,
) -> LstmResult:
    synthesis = request.synthesis
    model_params = request.model
    training = request.training
    device = _resolve_device(training.device)

    rng = np.random.default_rng(synthesis.seed)
    torch.manual_seed(synthesis.seed)
    if device.type == "cuda":
        torch.cuda.manual_seed_all(synthesis.seed)

    observed_time = np.arange(synthesis.n_samples) * synthesis.time_step
    observed_noise = rng.normal(0, synthesis.noise_std, synthesis.n_samples)
    observed = _series_values(observed_time, request, observed_noise)
    split = int(synthesis.n_samples * training.train_ratio)
    data_min = float(observed[:split].min())
    data_range = float(observed[:split].max() - data_min)
    if data_range <= np.finfo(np.float32).eps:
        data_range = 1.0
    normalized_data = (observed - data_min) / data_range
    time_scale = float(observed_time[split - 1])
    features = _build_features(normalized_data, observed_time, time_scale)

    train_dataset = TimeSeriesDataset(
        features[:split],
        normalized_data[:split],
        model_params.sequence_length,
        request.forecast_horizon,
    )
    validation_start = split - model_params.sequence_length
    validation_dataset = TimeSeriesDataset(
        features[validation_start:],
        normalized_data[validation_start:],
        model_params.sequence_length,
        request.forecast_horizon,
    )
    generator = torch.Generator().manual_seed(synthesis.seed)
    train_loader = DataLoader(
        train_dataset,
        batch_size=training.batch_size,
        shuffle=True,
        generator=generator,
    )
    validation_loader = DataLoader(
        validation_dataset,
        batch_size=training.batch_size,
        shuffle=False,
    )

    model = LstmPredictor(
        input_size=features.shape[1],
        hidden_size=model_params.hidden_size,
        num_layers=model_params.num_layers,
        dropout=model_params.dropout,
        future_feature_size=features.shape[1] - 1,
    ).to(device)
    training_started = time_module.perf_counter()
    train_losses, validation_losses = _train_model(
        model,
        train_loader,
        validation_loader,
        request,
        device,
        progress_callback,
    )
    training_seconds = time_module.perf_counter() - training_started

    if progress_callback is not None:
        progress_callback({"type": "predicting", "message": "正在预测中"})

    future_time = observed_time[-1] + synthesis.time_step * np.arange(
        1, request.forecast_horizon + 1
    )
    forecast_normalized = _predict_future(
        model,
        features,
        normalized_data,
        future_time,
        time_scale,
        model_params.sequence_length,
        device,
    )
    forecast = forecast_normalized * data_range + data_min
    future_noise = rng.normal(0, synthesis.noise_std, request.forecast_horizon)
    expected = _series_values(future_time, request, future_noise)

    errors = forecast - expected
    metrics = LstmTrainingMetrics(
        final_train_loss=train_losses[-1],
        final_validation_loss=validation_losses[-1],
        best_validation_loss=min(validation_losses),
        forecast_mae=float(np.mean(np.abs(errors))),
        forecast_rmse=float(np.sqrt(np.mean(np.square(errors)))),
        training_seconds=training_seconds,
        device=str(device),
    )
    result = LstmResult(
        user_id=user_id,
        dataset_params={
            **request.synthesis.model_dump(mode="json"),
            "forecast_horizon": request.forecast_horizon,
        },
        model_params=request.model.model_dump(mode="json"),
        training_params=request.training.model_dump(mode="json"),
        metrics=metrics.model_dump(mode="json"),
        train_losses=[float(value) for value in train_losses],
        validation_losses=[float(value) for value in validation_losses],
        observed_time=observed_time.astype(float).tolist(),
        observed_data=observed.astype(float).tolist(),
        future_time=future_time.astype(float).tolist(),
        forecast_data=forecast.astype(float).tolist(),
        expected_data=expected.astype(float).tolist(),
        create_time=datetime.now(UTC),
    )
    return lstmRepo.insert_result(result)


async def train_and_store(
    request: LstmTrainRequest,
    user_id: uuid.UUID,
) -> LstmResult:
    async with _training_semaphore:
        return await asyncio.to_thread(_train_and_store_sync, request, user_id)


async def train_and_store_events(
    request: LstmTrainRequest,
    user_id: uuid.UUID,
) -> AsyncIterator[dict[str, object]]:
    async with _training_semaphore:
        loop = asyncio.get_running_loop()
        event_queue: asyncio.Queue[dict[str, object]] = asyncio.Queue()

        def publish_progress(event: dict[str, object]) -> None:
            loop.call_soon_threadsafe(event_queue.put_nowait, event)

        def run_training() -> LstmResult:
            try:
                return _train_and_store_sync(request, user_id, publish_progress)
            finally:
                finished_event: dict[str, object] = {"type": "worker_finished"}
                loop.call_soon_threadsafe(
                    event_queue.put_nowait,
                    finished_event,
                )

        worker = asyncio.create_task(asyncio.to_thread(run_training))
        try:
            while True:
                event = await event_queue.get()
                if event["type"] == "worker_finished":
                    try:
                        result = await worker
                    except (ValueError, RuntimeError, SQLAlchemyError) as exc:
                        message = (
                            str(exc)
                            if isinstance(exc, ValueError)
                            else "训练失败，请稍后重试"
                        )
                        yield {"type": "error", "message": message}
                    else:
                        yield {"type": "completed", "result": result}
                    break

                yield event
        finally:
            if not worker.done():
                await asyncio.shield(worker)
