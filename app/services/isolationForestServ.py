import asyncio
import csv
import io
import time
import uuid
from datetime import UTC, datetime

import numpy as np
from matplotlib.figure import Figure
from openpyxl import Workbook
from sklearn.ensemble import IsolationForest

from app.models.isolation_forest import (
    IsolationForestDetectRequest,
    IsolationForestMetrics,
)
from app.models.tables.databaseTables import IsolationForestResult
from app.repositories import isolationForestRepo
from app.utils.plot import get_chinese_font

_detection_semaphore = asyncio.Semaphore(2)

_METRIC_LABELS = {
    "precision": "精确率(precision)",
    "recall": "召回率(recall)",
    "f1_score": "F1分数(f1_score)",
    "accuracy": "准确率(accuracy)",
    "detected_anomalies": "检测异常数(detected_anomalies)",
    "actual_anomalies": "实际异常数(actual_anomalies)",
    "detection_seconds": "检测耗时秒(detection_seconds)",
    "device": "计算设备(device)",
}

_ISOLATION_FOREST_PARAMETER_LABELS = {
    "normal_samples": "正常样本数(normal_samples)",
    "anomaly_samples": "异常样本数(anomaly_samples)",
    "cluster_std": "样本簇标准差(cluster_std)",
    "anomaly_radius_min": "异常点最小半径(anomaly_radius_min)",
    "anomaly_radius_max": "异常点最大半径(anomaly_radius_max)",
    "seed": "随机种子(seed)",
    "n_estimators": "孤立树数量(n_estimators)",
    "contamination": "预期异常比例(contamination)",
    "max_samples": "单树采样数(max_samples)",
}


def _generate_dataset(
    request: IsolationForestDetectRequest,
) -> tuple[np.ndarray, np.ndarray]:
    params = request.synthesis
    rng = np.random.default_rng(params.seed)
    first_cluster_size = params.normal_samples // 2
    normal = np.vstack(
        (
            rng.normal(
                loc=(-2.0, 0.0),
                scale=params.cluster_std,
                size=(first_cluster_size, 2),
            ),
            rng.normal(
                loc=(2.0, 0.0),
                scale=params.cluster_std,
                size=(params.normal_samples - first_cluster_size, 2),
            ),
        )
    )
    angles = rng.uniform(0, 2 * np.pi, params.anomaly_samples)
    radii = rng.uniform(
        params.anomaly_radius_min,
        params.anomaly_radius_max,
        params.anomaly_samples,
    )
    anomalies = np.column_stack((radii * np.cos(angles), radii * np.sin(angles)))

    values = np.vstack((normal, anomalies))
    actual_labels = np.concatenate(
        (
            np.zeros(params.normal_samples, dtype=int),
            np.ones(params.anomaly_samples, dtype=int),
        )
    )
    order = rng.permutation(len(values))
    return values[order], actual_labels[order]


def _calculate_metrics(
    actual_labels: np.ndarray,
    predicted_labels: np.ndarray,
    detection_seconds: float,
) -> IsolationForestMetrics:
    true_positive = int(np.sum((actual_labels == 1) & (predicted_labels == 1)))
    false_positive = int(np.sum((actual_labels == 0) & (predicted_labels == 1)))
    false_negative = int(np.sum((actual_labels == 1) & (predicted_labels == 0)))
    true_negative = int(np.sum((actual_labels == 0) & (predicted_labels == 0)))
    precision = true_positive / max(true_positive + false_positive, 1)
    recall = true_positive / max(true_positive + false_negative, 1)
    f1_score = 2 * precision * recall / max(precision + recall, np.finfo(float).eps)

    return IsolationForestMetrics(
        precision=precision,
        recall=recall,
        f1_score=f1_score,
        accuracy=(true_positive + true_negative) / len(actual_labels),
        detected_anomalies=int(predicted_labels.sum()),
        actual_anomalies=int(actual_labels.sum()),
        detection_seconds=detection_seconds,
    )


def _detect_and_store_sync(
    request: IsolationForestDetectRequest,
    user_id: uuid.UUID,
) -> IsolationForestResult:
    values, actual_labels = _generate_dataset(request)
    model_params = request.model
    model = IsolationForest(
        n_estimators=model_params.n_estimators,
        contamination=model_params.contamination,
        max_samples=model_params.max_samples,
        random_state=request.synthesis.seed,
        n_jobs=1,
    )
    detection_started = time.perf_counter()
    raw_predictions = model.fit_predict(values)
    anomaly_scores = -model.decision_function(values)
    detection_seconds = time.perf_counter() - detection_started
    predicted_labels = (raw_predictions == -1).astype(int)
    metrics = _calculate_metrics(
        actual_labels,
        predicted_labels,
        detection_seconds,
    )

    result = IsolationForestResult(
        user_id=user_id,
        dataset_params=request.synthesis.model_dump(mode="json"),
        model_params=request.model.model_dump(mode="json"),
        metrics=metrics.model_dump(mode="json"),
        x_values=values[:, 0].astype(float).tolist(),
        y_values=values[:, 1].astype(float).tolist(),
        actual_labels=actual_labels.astype(int).tolist(),
        predicted_labels=predicted_labels.astype(int).tolist(),
        anomaly_scores=anomaly_scores.astype(float).tolist(),
        create_time=datetime.now(UTC),
    )
    return isolationForestRepo.insert_result(result)


async def detect_and_store(
    request: IsolationForestDetectRequest,
    user_id: uuid.UUID,
) -> IsolationForestResult:
    async with _detection_semaphore:
        return await asyncio.to_thread(_detect_and_store_sync, request, user_id)


def get_result(
    result_id: uuid.UUID,
    user_id: uuid.UUID,
) -> IsolationForestResult | None:
    return isolationForestRepo.select_result_by_id_and_user_id(result_id, user_id)


def render_result_image(result: IsolationForestResult) -> bytes:
    chinese_font = get_chinese_font()
    x_values = np.asarray(result.x_values, dtype=float)
    y_values = np.asarray(result.y_values, dtype=float)
    actual_labels = np.asarray(result.actual_labels, dtype=bool)
    predicted_labels = np.asarray(result.predicted_labels, dtype=bool)
    true_positive = actual_labels & predicted_labels
    false_positive = ~actual_labels & predicted_labels
    false_negative = actual_labels & ~predicted_labels
    predicted_normal = ~actual_labels & ~predicted_labels

    figure = Figure(figsize=(14, 6), constrained_layout=True, facecolor="#f4f6f8")
    actual_axis, predicted_axis = figure.subplots(1, 2)
    precision = float(result.metrics["precision"])
    recall = float(result.metrics["recall"])
    figure.suptitle(
        f"孤立森林异常检测结果（精确率 {precision:.1%}，召回率 {recall:.1%}）",
        fontproperties=chinese_font,
        fontsize=17,
        fontweight="bold",
    )

    actual_axis.scatter(
        x_values[~actual_labels],
        y_values[~actual_labels],
        s=22,
        alpha=0.65,
        color="#2563eb",
        label="正常样本",
    )
    actual_axis.scatter(
        x_values[actual_labels],
        y_values[actual_labels],
        s=55,
        marker="x",
        linewidths=1.8,
        color="#dc2626",
        label="注入异常样本",
    )

    predicted_axis.scatter(
        x_values[predicted_normal],
        y_values[predicted_normal],
        s=22,
        alpha=0.65,
        color="#2563eb",
        label="正确识别的正常样本",
    )
    predicted_axis.scatter(
        x_values[true_positive],
        y_values[true_positive],
        s=55,
        marker="x",
        linewidths=1.8,
        color="#dc2626",
        label="正确检测的异常样本",
    )
    predicted_axis.scatter(
        x_values[false_positive],
        y_values[false_positive],
        s=55,
        marker="^",
        color="#9333ea",
        label="误报样本",
    )
    predicted_axis.scatter(
        x_values[false_negative],
        y_values[false_negative],
        s=70,
        facecolors="none",
        edgecolors="#f59e0b",
        linewidths=1.8,
        label="漏报样本",
    )

    for axis, title in (
        (actual_axis, "数据真实分布"),
        (predicted_axis, "模型检测结果"),
    ):
        axis.set_title(title, fontproperties=chinese_font, fontsize=13)
        axis.set_xlabel("特征一", fontproperties=chinese_font)
        axis.set_ylabel("特征二", fontproperties=chinese_font)
        axis.set_facecolor("white")
        axis.grid(alpha=0.2, linestyle="--")
        axis.legend(prop=chinese_font, frameon=False)
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)

    x_margin = max(float(np.ptp(x_values)) * 0.05, 0.5)
    y_margin = max(float(np.ptp(y_values)) * 0.05, 0.5)
    for axis in (actual_axis, predicted_axis):
        axis.set_xlim(
            float(x_values.min()) - x_margin, float(x_values.max()) + x_margin
        )
        axis.set_ylim(
            float(y_values.min()) - y_margin, float(y_values.max()) + y_margin
        )

    output = io.BytesIO()
    figure.savefig(output, format="png", dpi=150)
    return output.getvalue()


def export_result_csv(result: IsolationForestResult) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        (
            "样本(sample)",
            "特征一(feature_1)",
            "特征二(feature_2)",
            "实际异常(actual_anomaly)",
            "检测异常(detected_anomaly)",
            "异常分数(anomaly_score)",
        )
    )
    rows = zip(
        result.x_values,
        result.y_values,
        result.actual_labels,
        result.predicted_labels,
        result.anomaly_scores,
        strict=True,
    )
    for sample, values in enumerate(rows):
        writer.writerow((sample, *values))
    return output.getvalue().encode("utf-8-sig")


def export_result_excel(result: IsolationForestResult) -> bytes:
    workbook = Workbook()
    points_sheet = workbook.active
    if points_sheet is None:
        points_sheet = workbook.create_sheet()
    points_sheet.title = "数据点(points)"
    points_sheet.append(
        (
            "样本(sample)",
            "特征一(feature_1)",
            "特征二(feature_2)",
            "实际异常(actual_anomaly)",
            "检测异常(detected_anomaly)",
            "异常分数(anomaly_score)",
        )
    )
    rows = zip(
        result.x_values,
        result.y_values,
        result.actual_labels,
        result.predicted_labels,
        result.anomaly_scores,
        strict=True,
    )
    for sample, values in enumerate(rows):
        points_sheet.append((sample, *values))

    metrics_sheet = workbook.create_sheet("指标(metrics)")
    metrics_sheet.append(("名称(name)", "数值(value)"))
    for name, value in result.metrics.items():
        metrics_sheet.append((_METRIC_LABELS.get(name, name), value))

    params_sheet = workbook.create_sheet("参数(parameters)")
    params_sheet.append(("分组(group)", "参数名(name)", "参数值(value)"))
    for group, parameters in (
        ("数据集(dataset)", result.dataset_params),
        ("模型(model)", result.model_params),
    ):
        for name, value in parameters.items():
            params_sheet.append(
                (group, _ISOLATION_FOREST_PARAMETER_LABELS.get(name, name), value)
            )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
