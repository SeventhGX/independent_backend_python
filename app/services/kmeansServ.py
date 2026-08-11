import asyncio
import csv
import io
import time
import uuid
from datetime import UTC, datetime

import numpy as np
from matplotlib import colormaps
from matplotlib.figure import Figure
from openpyxl import Workbook
from sklearn.cluster import KMeans
from sklearn.metrics import (
    adjusted_rand_score,
    calinski_harabasz_score,
    davies_bouldin_score,
    silhouette_score,
)

from app.models.kmeans import KMeansClusterRequest, KMeansMetrics
from app.models.tables.databaseTables import KMeansResult
from app.repositories import kmeansRepo
from app.utils.plot import get_chinese_font

_clustering_semaphore = asyncio.Semaphore(2)

# 轮廓系数是 O(n^2) 计算，样本量大时改用抽样以保证 CPU 服务器上的响应速度。
_SILHOUETTE_SAMPLE_LIMIT = 2000

_METRIC_LABELS = {
    "inertia": "簇内平方和(inertia)",
    "silhouette": "轮廓系数(silhouette)",
    "davies_bouldin": "戴维森堡丁指数(davies_bouldin)",
    "calinski_harabasz": "方差比准则(calinski_harabasz)",
    "adjusted_rand_index": "调整兰德指数(adjusted_rand_index)",
    "iterations": "迭代次数(iterations)",
    "actual_k": "实际聚类簇数(actual_k)",
    "suggested_k": "推荐聚类簇数(suggested_k)",
    "sample_count": "样本总数(sample_count)",
    "cluster_seconds": "聚类耗时秒(cluster_seconds)",
    "device": "计算设备(device)",
}

_KMEANS_PARAMETER_LABELS = {
    "cluster_count": "真实簇数量(cluster_count)",
    "samples_per_cluster": "每簇样本数(samples_per_cluster)",
    "cluster_std": "簇标准差(cluster_std)",
    "center_spread": "簇中心分布半径(center_spread)",
    "seed": "随机种子(seed)",
    "n_clusters": "聚类簇数(n_clusters)",
    "init": "初始化方式(init)",
    "n_init": "重复初始化次数(n_init)",
    "max_iter": "最大迭代次数(max_iter)",
    "tol": "收敛阈值(tol)",
    "elbow_max_k": "肘部法则最大簇数(elbow_max_k)",
}


def _generate_dataset(
    request: KMeansClusterRequest,
) -> tuple[np.ndarray, np.ndarray]:
    params = request.synthesis
    rng = np.random.default_rng(params.seed)
    angles = np.linspace(0, 2 * np.pi, params.cluster_count, endpoint=False)
    centers = np.column_stack((np.cos(angles), np.sin(angles))) * params.center_spread
    centers += rng.normal(0, params.cluster_std * 0.3, centers.shape)
    values = np.vstack(
        [rng.normal(center, params.cluster_std, (params.samples_per_cluster, 2)) for center in centers]
    )
    true_labels = np.repeat(
        np.arange(params.cluster_count),
        params.samples_per_cluster,
    )
    order = rng.permutation(len(values))
    return values[order], true_labels[order]


def _silhouette(values: np.ndarray, labels: np.ndarray, seed: int) -> float:
    return float(
        silhouette_score(
            values,
            labels,
            sample_size=min(len(values), _SILHOUETTE_SAMPLE_LIMIT),
            random_state=seed,
        )
    )


def _build_model(request: KMeansClusterRequest, n_clusters: int) -> KMeans:
    model_params = request.model
    return KMeans(
        n_clusters=n_clusters,
        init=model_params.init,
        n_init=model_params.n_init,
        max_iter=model_params.max_iter,
        tol=model_params.tol,
        random_state=request.synthesis.seed,
    )


def _sweep_cluster_counts(
    values: np.ndarray,
    request: KMeansClusterRequest,
) -> tuple[list[int], list[float], list[float]]:
    """遍历候选簇数，产出肘部法则曲线与轮廓系数曲线。"""
    k_values: list[int] = []
    inertias: list[float] = []
    silhouettes: list[float] = []
    for n_clusters in range(2, request.evaluation.elbow_max_k + 1):
        model = _build_model(request, n_clusters)
        labels = model.fit_predict(values)
        k_values.append(n_clusters)
        inertias.append(float(model.inertia_))
        silhouettes.append(_silhouette(values, labels, request.synthesis.seed))
    return k_values, inertias, silhouettes


def _calculate_metrics(
    values: np.ndarray,
    true_labels: np.ndarray,
    cluster_labels: np.ndarray,
    model: KMeans,
    suggested_k: int,
    cluster_seconds: float,
    seed: int,
) -> KMeansMetrics:
    return KMeansMetrics(
        inertia=float(model.inertia_),
        silhouette=_silhouette(values, cluster_labels, seed),
        davies_bouldin=float(davies_bouldin_score(values, cluster_labels)),
        calinski_harabasz=float(calinski_harabasz_score(values, cluster_labels)),
        adjusted_rand_index=float(adjusted_rand_score(true_labels, cluster_labels)),
        iterations=int(model.n_iter_),
        actual_k=int(model.cluster_centers_.shape[0]),
        suggested_k=suggested_k,
        sample_count=len(values),
        cluster_seconds=cluster_seconds,
    )


def _cluster_and_store_sync(
    request: KMeansClusterRequest,
    user_id: uuid.UUID,
) -> KMeansResult:
    values, true_labels = _generate_dataset(request)
    model = _build_model(request, request.model.n_clusters)
    cluster_started = time.perf_counter()
    cluster_labels = model.fit_predict(values)
    cluster_seconds = time.perf_counter() - cluster_started
    center_distances = model.transform(values).min(axis=1)
    k_values, inertias, silhouettes = _sweep_cluster_counts(values, request)
    suggested_k = k_values[int(np.argmax(silhouettes))]
    metrics = _calculate_metrics(
        values,
        true_labels,
        cluster_labels,
        model,
        suggested_k,
        cluster_seconds,
        request.synthesis.seed,
    )

    result = KMeansResult(
        user_id=user_id,
        dataset_params=request.synthesis.model_dump(mode="json"),
        model_params=request.model.model_dump(mode="json"),
        evaluation_params=request.evaluation.model_dump(mode="json"),
        metrics=metrics.model_dump(mode="json"),
        x_values=values[:, 0].astype(float).tolist(),
        y_values=values[:, 1].astype(float).tolist(),
        true_labels=true_labels.astype(int).tolist(),
        cluster_labels=cluster_labels.astype(int).tolist(),
        center_distances=center_distances.astype(float).tolist(),
        center_x_values=model.cluster_centers_[:, 0].astype(float).tolist(),
        center_y_values=model.cluster_centers_[:, 1].astype(float).tolist(),
        elbow_k_values=k_values,
        elbow_inertias=inertias,
        elbow_silhouettes=silhouettes,
        create_time=datetime.now(UTC),
    )
    return kmeansRepo.insert_result(result)


async def cluster_and_store(
    request: KMeansClusterRequest,
    user_id: uuid.UUID,
) -> KMeansResult:
    async with _clustering_semaphore:
        return await asyncio.to_thread(_cluster_and_store_sync, request, user_id)


def get_result(
    result_id: uuid.UUID,
    user_id: uuid.UUID,
) -> KMeansResult | None:
    return kmeansRepo.select_result_by_id_and_user_id(result_id, user_id)


def _draw_clusters(axis, x_values, y_values, labels, label_prefix, chinese_font):
    palette = colormaps["tab10"]
    for label in np.unique(labels):
        mask = labels == label
        axis.scatter(
            x_values[mask],
            y_values[mask],
            s=20,
            alpha=0.7,
            color=palette(int(label) % 10),
            label=f"{label_prefix} {label}",
        )
    axis.legend(prop=chinese_font, frameon=False, ncol=2, fontsize=9)


def render_result_image(result: KMeansResult) -> bytes:
    chinese_font = get_chinese_font()
    x_values = np.asarray(result.x_values, dtype=float)
    y_values = np.asarray(result.y_values, dtype=float)
    true_labels = np.asarray(result.true_labels, dtype=int)
    cluster_labels = np.asarray(result.cluster_labels, dtype=int)
    center_x_values = np.asarray(result.center_x_values, dtype=float)
    center_y_values = np.asarray(result.center_y_values, dtype=float)
    k_values = np.asarray(result.elbow_k_values, dtype=int)
    inertias = np.asarray(result.elbow_inertias, dtype=float)
    silhouettes = np.asarray(result.elbow_silhouettes, dtype=float)

    figure = Figure(figsize=(19, 6), constrained_layout=True, facecolor="#f4f6f8")
    truth_axis, cluster_axis, elbow_axis = figure.subplots(1, 3)
    silhouette = float(result.metrics["silhouette"])
    adjusted_rand_index = float(result.metrics["adjusted_rand_index"])
    suggested_k = int(result.metrics["suggested_k"])
    figure.suptitle(
        f"K-Means 聚类结果（轮廓系数 {silhouette:.3f}，调整兰德指数 {adjusted_rand_index:.3f}）",
        fontproperties=chinese_font,
        fontsize=17,
        fontweight="bold",
    )

    _draw_clusters(truth_axis, x_values, y_values, true_labels, "真实簇", chinese_font)
    _draw_clusters(cluster_axis, x_values, y_values, cluster_labels, "聚类簇", chinese_font)
    cluster_axis.scatter(
        center_x_values,
        center_y_values,
        s=180,
        marker="X",
        color="#111827",
        edgecolors="white",
        linewidths=1.5,
        zorder=3,
        label="聚类质心",
    )
    cluster_axis.legend(prop=chinese_font, frameon=False, ncol=2, fontsize=9)

    for axis, title in (
        (truth_axis, "数据真实分布"),
        (cluster_axis, "K-Means 聚类结果"),
    ):
        axis.set_title(title, fontproperties=chinese_font, fontsize=13)
        axis.set_xlabel("特征一", fontproperties=chinese_font)
        axis.set_ylabel("特征二", fontproperties=chinese_font)
        axis.set_facecolor("white")
        axis.grid(alpha=0.2, linestyle="--")
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)

    silhouette_axis = elbow_axis.twinx()
    inertia_line = elbow_axis.plot(
        k_values,
        inertias,
        marker="o",
        color="#2563eb",
        label="簇内平方和(inertia)",
    )
    silhouette_line = silhouette_axis.plot(
        k_values,
        silhouettes,
        marker="s",
        color="#f59e0b",
        label="轮廓系数(silhouette)",
    )
    suggested_line = elbow_axis.axvline(
        suggested_k,
        color="#dc2626",
        linestyle="--",
        alpha=0.8,
        label=f"推荐簇数 K={suggested_k}",
    )
    elbow_axis.set_title("肘部法则与轮廓系数", fontproperties=chinese_font, fontsize=13)
    elbow_axis.set_xlabel("聚类簇数 K", fontproperties=chinese_font)
    elbow_axis.set_ylabel("簇内平方和", fontproperties=chinese_font, color="#2563eb")
    silhouette_axis.set_ylabel("轮廓系数", fontproperties=chinese_font, color="#f59e0b")
    elbow_axis.set_xticks(k_values)
    elbow_axis.set_facecolor("white")
    elbow_axis.grid(alpha=0.2, linestyle="--")
    elbow_axis.spines["top"].set_visible(False)
    silhouette_axis.spines["top"].set_visible(False)
    elbow_handles = [*inertia_line, *silhouette_line, suggested_line]
    elbow_axis.legend(
        handles=elbow_handles,
        labels=[handle.get_label() for handle in elbow_handles],
        prop=chinese_font,
        frameon=False,
        fontsize=9,
    )

    for axis in (truth_axis, cluster_axis):
        x_margin = max(float(np.ptp(x_values)) * 0.05, 0.5)
        y_margin = max(float(np.ptp(y_values)) * 0.05, 0.5)
        axis.set_xlim(float(x_values.min()) - x_margin, float(x_values.max()) + x_margin)
        axis.set_ylim(float(y_values.min()) - y_margin, float(y_values.max()) + y_margin)

    output = io.BytesIO()
    figure.savefig(output, format="png", dpi=150)
    return output.getvalue()


_POINT_HEADERS = (
    "样本(sample)",
    "特征一(feature_1)",
    "特征二(feature_2)",
    "真实簇(true_cluster)",
    "聚类簇(cluster)",
    "到质心距离(center_distance)",
)


def _iterate_points(result: KMeansResult):
    rows = zip(
        result.x_values,
        result.y_values,
        result.true_labels,
        result.cluster_labels,
        result.center_distances,
        strict=True,
    )
    for sample, values in enumerate(rows):
        yield (sample, *values)


def export_result_csv(result: KMeansResult) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(_POINT_HEADERS)
    writer.writerows(_iterate_points(result))
    return output.getvalue().encode("utf-8-sig")


def export_result_excel(result: KMeansResult) -> bytes:
    workbook = Workbook()
    points_sheet = workbook.active
    if points_sheet is None:
        points_sheet = workbook.create_sheet()
    points_sheet.title = "数据点(points)"
    points_sheet.append(_POINT_HEADERS)
    for row in _iterate_points(result):
        points_sheet.append(row)

    centers_sheet = workbook.create_sheet("质心(centers)")
    centers_sheet.append(("聚类簇(cluster)", "特征一(feature_1)", "特征二(feature_2)"))
    for cluster, center in enumerate(zip(result.center_x_values, result.center_y_values, strict=True)):
        centers_sheet.append((cluster, *center))

    elbow_sheet = workbook.create_sheet("肘部法则(elbow)")
    elbow_sheet.append(("聚类簇数(k)", "簇内平方和(inertia)", "轮廓系数(silhouette)"))
    for row in zip(
        result.elbow_k_values,
        result.elbow_inertias,
        result.elbow_silhouettes,
        strict=True,
    ):
        elbow_sheet.append(row)

    metrics_sheet = workbook.create_sheet("指标(metrics)")
    metrics_sheet.append(("名称(name)", "数值(value)"))
    for name, value in result.metrics.items():
        metrics_sheet.append((_METRIC_LABELS.get(name, name), value))

    params_sheet = workbook.create_sheet("参数(parameters)")
    params_sheet.append(("分组(group)", "参数名(name)", "参数值(value)"))
    for group, parameters in (
        ("数据集(dataset)", result.dataset_params),
        ("模型(model)", result.model_params),
        ("评估(evaluation)", result.evaluation_params),
    ):
        for name, value in parameters.items():
            params_sheet.append((group, _KMEANS_PARAMETER_LABELS.get(name, name), value))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
