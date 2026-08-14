import asyncio
import csv
import heapq
import io
import itertools
import math
import time
import uuid
from collections import defaultdict
from datetime import UTC, datetime

from openpyxl import Workbook

from app.models.astar import (
    AStarMetrics,
    AStarNeighborEvaluation,
    AStarOpenNode,
    AStarSolveRequest,
    AStarStep,
)
from app.models.tables.databaseTables import AStarResult
from app.repositories import astarRepo

_solve_semaphore = asyncio.Semaphore(4)

_METRIC_LABELS = {
    "found": "是否找到路径(found)",
    "path_cost": "路径总代价(path_cost)",
    "path_euclidean_distance": "路径欧氏长度(path_euclidean_distance)",
    "heuristic_scale": "启发函数缩放系数(heuristic_scale)",
    "explored_nodes": "探索节点数(explored_nodes)",
    "step_count": "求解步数(step_count)",
    "node_count": "节点数(node_count)",
    "edge_count": "边数(edge_count)",
    "solve_seconds": "求解耗时秒(solve_seconds)",
}


def _normalize_graph(request: AStarSolveRequest) -> dict[str, object]:
    node_by_id = {node.id: node for node in request.nodes}
    used_edge_ids = {edge.id for edge in request.edges if edge.id is not None}
    edges = []
    for index, edge in enumerate(request.edges):
        source = node_by_id[edge.source]
        target = node_by_id[edge.target]
        weight = edge.weight
        if weight is None:
            weight = math.hypot(source.x - target.x, source.y - target.y)
        edge_id = edge.id or f"edge-{index}"
        while edge_id in used_edge_ids and edge.id is None:
            edge_id = f"{edge_id}-generated"
        used_edge_ids.add(edge_id)
        edges.append(
            {
                "id": edge_id,
                "source": edge.source,
                "target": edge.target,
                "weight": weight,
            }
        )
    return {
        "nodes": [node.model_dump(mode="json") for node in request.nodes],
        "edges": edges,
        "directed": request.directed,
    }


def _heuristic_scale(graph: dict[str, object], node_by_id: dict[str, dict]) -> float:
    scale = 1.0
    for edge in graph["edges"]:
        source = node_by_id[edge["source"]]
        target = node_by_id[edge["target"]]
        euclidean_distance = math.hypot(
            source["x"] - target["x"],
            source["y"] - target["y"],
        )
        if euclidean_distance > 0:
            scale = min(scale, edge["weight"] / euclidean_distance)
    return scale


def _build_open_set(
    g_scores: dict[str, float],
    parents: dict[str, str],
    closed: set[str],
    heuristic,
) -> list[AStarOpenNode]:
    nodes = [
        AStarOpenNode(
            node_id=node_id,
            g=g_score,
            h=heuristic(node_id),
            f=g_score + heuristic(node_id),
            parent_id=parents.get(node_id),
        )
        for node_id, g_score in g_scores.items()
        if node_id not in closed
    ]
    return sorted(nodes, key=lambda node: (node.f, node.h, node.node_id))


def _reconstruct_path(
    goal_id: str,
    parents: dict[str, str],
    parent_edges: dict[str, str],
) -> tuple[list[str], list[str]]:
    path = [goal_id]
    edge_ids = []
    while path[-1] in parents:
        edge_ids.append(parent_edges[path[-1]])
        path.append(parents[path[-1]])
    path.reverse()
    edge_ids.reverse()
    return path, edge_ids


def _path_euclidean_distance(path: list[str], node_by_id: dict[str, dict]) -> float:
    return sum(
        math.hypot(
            node_by_id[target]["x"] - node_by_id[source]["x"],
            node_by_id[target]["y"] - node_by_id[source]["y"],
        )
        for source, target in itertools.pairwise(path)
    )


def _solve_and_store_sync(
    request: AStarSolveRequest,
    user_id: uuid.UUID,
) -> AStarResult:
    graph = _normalize_graph(request)
    node_by_id = {node["id"]: node for node in graph["nodes"]}
    adjacency: dict[str, list[tuple[str, float, str]]] = defaultdict(list)
    for edge in graph["edges"]:
        adjacency[edge["source"]].append(
            (edge["target"], edge["weight"], edge["id"])
        )
        if not graph["directed"]:
            adjacency[edge["target"]].append(
                (edge["source"], edge["weight"], edge["id"])
            )
    for neighbors in adjacency.values():
        neighbors.sort(key=lambda item: (item[0], item[2]))

    goal = node_by_id[request.goal_id]
    heuristic_scale = _heuristic_scale(graph, node_by_id)

    def heuristic(node_id: str) -> float:
        node = node_by_id[node_id]
        return heuristic_scale * math.hypot(node["x"] - goal["x"], node["y"] - goal["y"])

    started_at = time.perf_counter()
    sequence = itertools.count()
    open_heap = [(heuristic(request.start_id), 0.0, next(sequence), request.start_id)]
    g_scores = {request.start_id: 0.0}
    parents: dict[str, str] = {}
    parent_edges: dict[str, str] = {}
    closed: set[str] = set()
    closed_order: list[str] = []
    steps: list[AStarStep] = []
    path: list[str] = []
    path_edge_ids: list[str] = []

    while open_heap:
        _, queued_g, _, current_id = heapq.heappop(open_heap)
        if current_id in closed or not math.isclose(queued_g, g_scores[current_id]):
            continue

        closed.add(current_id)
        closed_order.append(current_id)
        evaluations: list[AStarNeighborEvaluation] = []
        if current_id == request.goal_id:
            path, path_edge_ids = _reconstruct_path(
                current_id,
                parents,
                parent_edges,
            )
        else:
            for neighbor_id, weight, edge_id in adjacency[current_id]:
                tentative_g = g_scores[current_id] + weight
                previous_g = g_scores.get(neighbor_id)
                accepted = neighbor_id not in closed and (
                    previous_g is None or tentative_g < previous_g
                )
                evaluations.append(
                    AStarNeighborEvaluation(
                        edge_id=edge_id,
                        node_id=neighbor_id,
                        edge_weight=weight,
                        tentative_g=tentative_g,
                        previous_g=previous_g,
                        accepted=accepted,
                    )
                )
                if accepted:
                    g_scores[neighbor_id] = tentative_g
                    parents[neighbor_id] = current_id
                    parent_edges[neighbor_id] = edge_id
                    heapq.heappush(
                        open_heap,
                        (
                            tentative_g + heuristic(neighbor_id),
                            tentative_g,
                            next(sequence),
                            neighbor_id,
                        ),
                    )

        steps.append(
            AStarStep(
                iteration=len(steps),
                current_id=current_id,
                open_set=_build_open_set(g_scores, parents, closed, heuristic),
                closed_set=closed_order.copy(),
                neighbors=evaluations,
            )
        )
        if path:
            break

    solve_seconds = time.perf_counter() - started_at
    start = node_by_id[request.start_id]
    start_to_goal_distance = math.hypot(
        start["x"] - goal["x"],
        start["y"] - goal["y"],
    )
    metrics = AStarMetrics(
        found=bool(path),
        path_cost=g_scores.get(request.goal_id) if path else None,
        path_euclidean_distance=(
            _path_euclidean_distance(path, node_by_id) if path else None
        ),
        heuristic_scale=heuristic_scale,
        explored_nodes=len(closed),
        step_count=len(steps),
        node_count=len(request.nodes),
        edge_count=len(request.edges),
        solve_seconds=solve_seconds,
    )
    result = AStarResult(
        user_id=user_id,
        graph=graph,
        start_id=request.start_id,
        goal_id=request.goal_id,
        start_to_goal_distance=start_to_goal_distance,
        path=path,
        path_edge_ids=path_edge_ids,
        steps=[step.model_dump(mode="json") for step in steps],
        metrics=metrics.model_dump(mode="json"),
        create_time=datetime.now(UTC),
    )
    return astarRepo.insert_result(result)


async def solve_and_store(
    request: AStarSolveRequest,
    user_id: uuid.UUID,
) -> AStarResult:
    async with _solve_semaphore:
        return await asyncio.to_thread(_solve_and_store_sync, request, user_id)


def get_result(result_id: uuid.UUID, user_id: uuid.UUID) -> AStarResult | None:
    return astarRepo.select_result_by_id_and_user_id(result_id, user_id)


def _write_csv_section(writer, title: str, headers: tuple[str, ...], rows) -> None:
    writer.writerow((title,))
    writer.writerow(headers)
    writer.writerows(rows)
    writer.writerow(())


def export_result_csv(result: AStarResult) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    _write_csv_section(
        writer,
        "节点(nodes)",
        ("节点ID(node_id)", "标签(label)", "横坐标(x)", "纵坐标(y)"),
        ((node["id"], node.get("label"), node["x"], node["y"]) for node in result.graph["nodes"]),
    )
    _write_csv_section(
        writer,
        "边(edges)",
        ("边ID(edge_id)", "起点(source)", "终点(target)", "权重(weight)"),
        ((edge["id"], edge["source"], edge["target"], edge["weight"]) for edge in result.graph["edges"]),
    )
    _write_csv_section(
        writer,
        "路径(path)",
        ("顺序(order)", "节点ID(node_id)", "到达边ID(edge_id)"),
        (
            (index, node_id, result.path_edge_ids[index - 1] if index else None)
            for index, node_id in enumerate(result.path)
        ),
    )
    _write_csv_section(
        writer,
        "求解步骤(steps)",
        ("步骤(iteration)", "当前节点(current)", "开放集合(open_set)", "关闭集合(closed_set)", "邻边评估(neighbors)"),
        (
            (
                step["iteration"],
                step["current_id"],
                ";".join(node["node_id"] for node in step["open_set"]),
                ";".join(step["closed_set"]),
                ";".join(
                    f'{item["edge_id"]}:{item["node_id"]}:{"accepted" if item["accepted"] else "rejected"}'
                    for item in step["neighbors"]
                ),
            )
            for step in result.steps
        ),
    )
    _write_csv_section(
        writer,
        "指标(metrics)",
        ("名称(name)", "数值(value)"),
        ((_METRIC_LABELS.get(name, name), value) for name, value in result.metrics.items()),
    )
    return output.getvalue().encode("utf-8-sig")


def export_result_excel(result: AStarResult) -> bytes:
    workbook = Workbook()
    nodes_sheet = workbook.active
    if nodes_sheet is None:
        nodes_sheet = workbook.create_sheet()
    nodes_sheet.title = "节点(nodes)"
    nodes_sheet.append(("节点ID(node_id)", "标签(label)", "横坐标(x)", "纵坐标(y)"))
    for node in result.graph["nodes"]:
        nodes_sheet.append((node["id"], node.get("label"), node["x"], node["y"]))

    edges_sheet = workbook.create_sheet("边(edges)")
    edges_sheet.append(("边ID(edge_id)", "起点(source)", "终点(target)", "权重(weight)"))
    for edge in result.graph["edges"]:
        edges_sheet.append((edge["id"], edge["source"], edge["target"], edge["weight"]))

    path_sheet = workbook.create_sheet("路径(path)")
    path_sheet.append(("顺序(order)", "节点ID(node_id)", "到达边ID(edge_id)"))
    for index, node_id in enumerate(result.path):
        path_sheet.append((index, node_id, result.path_edge_ids[index - 1] if index else None))

    steps_sheet = workbook.create_sheet("求解步骤(steps)")
    steps_sheet.append(("步骤(iteration)", "当前节点(current)", "开放集合(open_set)", "关闭集合(closed_set)"))
    for step in result.steps:
        steps_sheet.append(
            (
                step["iteration"],
                step["current_id"],
                ";".join(node["node_id"] for node in step["open_set"]),
                ";".join(step["closed_set"]),
            )
        )

    evaluations_sheet = workbook.create_sheet("邻边评估(evaluations)")
    evaluations_sheet.append(
        (
            "步骤(iteration)",
            "边ID(edge_id)",
            "相邻节点(node_id)",
            "边权(edge_weight)",
            "候选代价(tentative_g)",
            "原代价(previous_g)",
            "是否接受(accepted)",
        )
    )
    for step in result.steps:
        for item in step["neighbors"]:
            evaluations_sheet.append(
                (
                    step["iteration"],
                    item["edge_id"],
                    item["node_id"],
                    item["edge_weight"],
                    item["tentative_g"],
                    item["previous_g"],
                    item["accepted"],
                )
            )

    metrics_sheet = workbook.create_sheet("指标(metrics)")
    metrics_sheet.append(("名称(name)", "数值(value)"))
    for name, value in result.metrics.items():
        metrics_sheet.append((_METRIC_LABELS.get(name, name), value))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()